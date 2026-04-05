#!/usr/bin/env python3
"""
Excel Parser - Supports .xlsx, .xls, .csv formats
Uses openpyxl for .xlsx and xlrd for .xls
"""

import openpyxl
import sys
import os
import csv
import json
import zipfile
import re
import xml.etree.ElementTree as ET

def parse_excel(file_path):
    """Parse Excel/CSV file and return sheets data as JSON"""
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        
        if file_ext == '.csv':
            return parse_csv(file_path)
        elif file_ext == '.xls':
            return parse_xls(file_path)
        else:
            # .xlsx, .xlsm, .xltx, .xltm
            return parse_xlsx(file_path)
    
    except Exception as e:
        import traceback
        return {"ok": False, "error": str(e), "traceback": traceback.format_exc()}

def parse_csv(file_path):
    """Parse CSV file"""
    try:
        sheets = {}
        rows = []
        
        encodings = ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']
        
        for encoding in encodings:
            try:
                with open(file_path, 'r', encoding=encoding, newline='') as f:
                    reader = csv.reader(f)
                    for row in reader:
                        row_data = [str(cell).strip() if cell is not None else "" for cell in row]
                        rows.append(row_data)
                break
            except UnicodeDecodeError:
                continue
        
        if not rows:
            rows = [[""]]
        
        sheets["Sheet1"] = rows
        
        return {"ok": True, "sheets": sheets, "names": ["Sheet1"]}
    except Exception as e:
        return {"ok": False, "error": f"CSV parse error: {str(e)}"}

def parse_xls(file_path):
    """Parse .xls files using xlrd"""
    try:
        import xlrd
        
        wb = xlrd.open_workbook(file_path)
        sheets = {}
        sheet_names = wb.sheet_names()
        
        for sheet_name in sheet_names:
            ws = wb.sheet_by_name(sheet_name)
            rows = []
            
            for row_idx in range(ws.nrows):
                row_data = []
                for col_idx in range(ws.ncols):
                    cell = ws.cell(row_idx, col_idx)
                    value = cell.value
                    
                    if value is None or value == "":
                        row_data.append("")
                    elif isinstance(value, float) and value == int(value):
                        row_data.append(str(int(value)))
                    else:
                        row_data.append(str(value).strip())
                
                rows.append(row_data)
            
            sheets[sheet_name] = rows if rows else [[""]]
        
        wb.release_resources()
        
        return {"ok": True, "sheets": sheets, "names": sheet_names}
    except ImportError:
        return {"ok": False, "error": "xlrd not installed. Run: pip3 install xlrd"}
    except Exception as e:
        return {"ok": False, "error": f"XLS parse error: {str(e)}"}

def extract_xlsx_data(file_path, data_only=True):
    """Extract data from XLSX file - helper function"""
    try:
        wb = openpyxl.load_workbook(file_path, data_only=data_only)
        sheets = {}
        sheet_names = wb.sheetnames
        
        for sheet_name in sheet_names:
            ws = wb[sheet_name]
            rows = []
            
            # Get the dimensions to know the actual used range
            if ws.dimensions:
                # Iterate through all rows in the used range
                for row in ws.iter_rows(values_only=True):
                    row_data = []
                    for cell_value in row:
                        if cell_value is None or cell_value == "":
                            row_data.append("")
                        elif isinstance(cell_value, float) and cell_value == int(cell_value):
                            row_data.append(str(int(cell_value)))
                        else:
                            row_data.append(str(cell_value).strip())
                    rows.append(row_data)
            
            # Remove trailing empty rows but keep at least one
            while len(rows) > 1 and all(cell == "" for cell in rows[-1]):
                rows.pop()
            
            if not rows:
                rows = [[""]]
            
            sheets[sheet_name] = rows
        
        wb.close()
        return sheets, sheet_names
        
    except Exception as e:
        raise Exception(f"Failed to extract XLSX data with data_only={data_only}: {str(e)}")

def _xlsx_sheet_drawing_map(zf):
    """Return mapping: sheet_name -> drawing_part_path (e.g. xl/drawings/drawing1.xml)"""
    # Read workbook to get sheet names + relationship ids
    wb_root = ET.fromstring(zf.read('xl/workbook.xml'))
    wb_rels_root = ET.fromstring(zf.read('xl/_rels/workbook.xml.rels'))

    ns = {
        'main': 'http://schemas.openxmlformats.org/spreadsheetml/2006/main',
        'r': 'http://schemas.openxmlformats.org/officeDocument/2006/relationships',
        'rels': 'http://schemas.openxmlformats.org/package/2006/relationships'
    }

    rid_to_target = {}
    for rel in wb_rels_root.findall('rels:Relationship', ns):
        rid = rel.attrib.get('Id')
        target = rel.attrib.get('Target')
        if rid and target:
            # Targets in workbook rels are relative to xl/
            rid_to_target[rid] = 'xl/' + target.lstrip('/')

    sheet_name_to_sheet_part = {}
    for sh in wb_root.findall('main:sheets/main:sheet', ns):
        name = sh.attrib.get('name')
        rid = sh.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
        if name and rid and rid in rid_to_target:
            sheet_name_to_sheet_part[name] = rid_to_target[rid]

    # For each worksheet part, resolve its drawing relationship
    out = {}
    for sheet_name, sheet_part in sheet_name_to_sheet_part.items():
        try:
            sheet_xml = ET.fromstring(zf.read(sheet_part))
        except KeyError:
            continue

        # Find <drawing r:id="rIdX"/>
        drawing_el = sheet_xml.find('main:drawing', ns)
        if drawing_el is None:
            continue

        drawing_rid = drawing_el.attrib.get('{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id')
        if not drawing_rid:
            continue

        # Worksheet rels file
        # sheet_part: xl/worksheets/sheet1.xml -> rels: xl/worksheets/_rels/sheet1.xml.rels
        rels_path = sheet_part.replace('xl/worksheets/', 'xl/worksheets/_rels/') + '.rels'
        try:
            rels_xml = ET.fromstring(zf.read(rels_path))
        except KeyError:
            continue

        drawing_target = None
        for rel in rels_xml.findall('rels:Relationship', ns):
            if rel.attrib.get('Id') == drawing_rid:
                drawing_target = rel.attrib.get('Target')
                break
        if not drawing_target:
            continue

        # Target is relative to xl/worksheets/
        # typically "../drawings/drawing1.xml"
        drawing_part = os.path.normpath('xl/worksheets/' + drawing_target).replace('xl/worksheets/xl/', 'xl/')
        if drawing_part.startswith('xl/'):
            out[sheet_name] = drawing_part
    return out

def _extract_text_from_drawing_xml(xml_bytes):
    """Extract anchored texts from a drawing part. Returns list of (row, col, text) where row/col are 0-based."""
    root = ET.fromstring(xml_bytes)
    ns = {
        'xdr': 'http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing',
        'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'
    }

    def collect_text(tx_body_el):
        if tx_body_el is None:
            return ''
        parts = []
        # paragraphs
        for p in tx_body_el.findall('.//a:p', ns):
            runs = []
            for t in p.findall('.//a:t', ns):
                if t.text:
                    runs.append(t.text)
            para = ''.join(runs).strip()
            if para:
                parts.append(para)
        return '\n'.join(parts).strip()

    anchored = []
    # twoCellAnchor / oneCellAnchor
    for anchor in root.findall('xdr:twoCellAnchor', ns) + root.findall('xdr:oneCellAnchor', ns):
        from_el = anchor.find('xdr:from', ns)
        if from_el is None:
            continue
        col_el = from_el.find('xdr:col', ns)
        row_el = from_el.find('xdr:row', ns)
        if col_el is None or row_el is None or col_el.text is None or row_el.text is None:
            continue
        try:
            col = int(col_el.text)
            row = int(row_el.text)
        except ValueError:
            continue

        # Shape text is usually in sp/txBody
        tx_body = anchor.find('.//xdr:sp//a:txBody', ns)
        text = collect_text(tx_body)
        if text:
            anchored.append((row, col, text))
    return anchored

def extract_xlsx_shapes_text(file_path):
    """Fallback: Extract text from shapes/textboxes inside XLSX drawings and place into a grid."""
    sheets = {}
    sheet_names = []
    with zipfile.ZipFile(file_path, 'r') as zf:
        sheet_drawing = _xlsx_sheet_drawing_map(zf)
        for sheet_name, drawing_part in sheet_drawing.items():
            sheet_names.append(sheet_name)
            try:
                anchored = _extract_text_from_drawing_xml(zf.read(drawing_part))
            except KeyError:
                sheets[sheet_name] = [[""]]
                continue

            if not anchored:
                sheets[sheet_name] = [[""]]
                continue

            max_row = max(r for r, _, _ in anchored)
            max_col = max(c for _, c, _ in anchored)
            # Safety caps
            max_row = min(max_row, 500)
            max_col = min(max_col, 200)
            grid = [["" for _ in range(max_col + 1)] for __ in range(max_row + 1)]
            for r, c, text in anchored:
                if r <= max_row and c <= max_col:
                    grid[r][c] = text
            sheets[sheet_name] = grid

    return sheets, sheet_names

def parse_xlsx(file_path):
    """Parse .xlsx files - FIXED to properly extract all data"""
    
    try:
        source = 'cells'
        # Try with data_only=True first (gets calculated values from formulas)
        sheets, sheet_names = extract_xlsx_data(file_path, data_only=True)
        
        # Check if we got any real data
        total_non_empty_cells = sum(
            sum(1 for cell in row if cell != "")
            for sheet in sheets.values()
            for row in sheet
        )
        
        # If no data found, fallback to data_only=False (gets raw values/formulas)
        if total_non_empty_cells == 0:
            sheets, sheet_names = extract_xlsx_data(file_path, data_only=False)

        # If still empty, try extracting textbox/shape text from drawings
        total_non_empty_cells = sum(
            sum(1 for cell in row if cell != "")
            for sheet in sheets.values()
            for row in sheet
        )
        if total_non_empty_cells == 0:
            shape_sheets, shape_names = extract_xlsx_shapes_text(file_path)
            shape_non_empty = sum(
                sum(1 for cell in row if cell != "")
                for sheet in shape_sheets.values()
                for row in sheet
            )
            if shape_non_empty > 0:
                sheets, sheet_names = shape_sheets, shape_names
                source = 'shapes'
        total_non_empty_cells = sum(
            sum(1 for cell in row if cell != "")
            for sheet in sheets.values()
            for row in sheet
        )
        return {"ok": True, "sheets": sheets, "names": sheet_names, "meta": {"source": source, "non_empty_cells": total_non_empty_cells}}
        
    except Exception as e:
        # Try fallback method with data_only=False
        try:
            sheets, sheet_names = extract_xlsx_data(file_path, data_only=False)
            total_non_empty_cells = sum(
                sum(1 for cell in row if cell != "")
                for sheet in sheets.values()
                for row in sheet
            )
            return {"ok": True, "sheets": sheets, "names": sheet_names, "meta": {"source": 'cells', "non_empty_cells": total_non_empty_cells}}
        except Exception as e2:
            import traceback
            return {"ok": False, "error": f"XLSX parse error: {str(e)}", "traceback": traceback.format_exc()}

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(json.dumps({"ok": False, "error": "No file path provided"}))
        sys.exit(1)
    
    file_path = sys.argv[1]
    
    if not os.path.exists(file_path):
        print(json.dumps({"ok": False, "error": f"File not found: {file_path}"}))
        sys.exit(1)
    
    result = parse_excel(file_path)
    print(json.dumps(result))
